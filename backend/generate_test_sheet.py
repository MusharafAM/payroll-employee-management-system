import openpyxl
from openpyxl.styles import Font, Alignment

def generate_sheet():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Timecard Report"

    # Set some metadata headers to satisfy parser auto-detection (which checks for "employee timecard" or "pay period")
    ws.cell(row=1, column=1, value="PAY PERIOD: 01-06-2026 to 30-06-2026")
    ws.cell(row=2, column=1, value="Employee Timecard Report")
    
    # Employee 1 Header (matches EMP-EMPL-2876 in database)
    ws.cell(row=4, column=1, value="Employee")
    ws.cell(row=4, column=4, value="Dev Employee (EMP-EMPL-2876)")
    
    # Row headers
    headers = ["Day", "Date", "Time In", "Time Out", "Shift", "Daily Total", "Note"]
    for col_idx, h in enumerate(headers, 1):
        ws.cell(row=5, column=col_idx, value=h)

    # Shift data for Employee 1 (June 1st to June 5th)
    data_emp1 = [
        ("MON", "01-06-26", "08:00", "17:00", "Day Shift", "09:00", ""),
        ("TUE", "02-06-26", "08:15", "17:15", "Day Shift", "09:00", ""),
        ("WED", "03-06-26", "08:00", "18:00", "Day Shift", "10:00", ""), # 10h total (8 reg + 1 OT after lunch deduction)
        ("THU", "04-06-26", "08:30", "17:30", "Day Shift", "09:00", ""),
        ("FRI", "05-06-26", "08:00", "17:00", "Day Shift", "09:00", "")
    ]
    
    current_row = 6
    for row in data_emp1:
        for col_idx, val in enumerate(row, 1):
            ws.cell(row=current_row, column=col_idx, value=val)
        current_row += 1

    # Add spacing
    current_row += 2

    # Employee 2 Header (matches EMP-musha-221 in database)
    ws.cell(row=current_row, column=1, value="Employee")
    ws.cell(row=current_row, column=4, value="Abdhul Munaf Musharaf (221)")
    current_row += 1
    
    for col_idx, h in enumerate(headers, 1):
        ws.cell(row=current_row, column=col_idx, value=h)
    current_row += 1

    # Shift data for Employee 2 (June 1st to June 5th)
    data_emp2 = [
        ("MON", "01-06-26", "08:00", "17:00", "Day Shift", "09:00", ""),
        ("TUE", "02-06-26", "08:00", "17:00", "Day Shift", "09:00", ""),
        ("WED", "03-06-26", "08:00", "17:00", "Day Shift", "09:00", ""),
        ("THU", "04-06-26", "08:00", "12:00", "Day Shift", "04:00", ""), # 4h total (3 reg after lunch deduction)
        ("FRI", "05-06-26", "08:00", "17:00", "Day Shift", "09:00", "")
    ]

    for row in data_emp2:
        for col_idx, val in enumerate(row, 1):
            ws.cell(row=current_row, column=col_idx, value=val)
        current_row += 1

    # Style cells slightly
    for col in ws.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        ws.column_dimensions[col[0].column_letter].width = max(max_len + 3, 10)

    # Save spreadsheet to workspace root
    output_path = "../sample_attendance.xlsx"
    wb.save(output_path)
    print(f"✅ Created sample biometric Excel spreadsheet at: sample_attendance.xlsx")

if __name__ == "__main__":
    generate_sheet()
